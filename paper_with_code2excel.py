import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font, RichTextProperties
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.descriptors.excel import ExtensionList, Extension
from tomlkit import datetime
import datetime
import json
import argparse
def split_ex(str,a_i='{',a_o='}',split=','):
    cnt=0
    result = []
    tmp_str = ""
    for s in str:
        if s == a_i:
            tmp_str += "{"
            cnt+=1
        elif s == a_o:
            cnt-=1
            tmp_str += "}"
        elif s == split:
            if cnt == 0:
                result.append(tmp_str)
                tmp_str = ""
            else:
                tmp_str += s
        else:
            tmp_str += s
    return result

def main():
    #url = "https://paperswithcode.com/sota/image-classification-on-imagenet"
    url = input("paper with code sota URL >> ")
    parser = argparse.ArgumentParser()
    parser.add_argument('url',type=str)
    parser.add_argument('--all_label',action='store_true')
    args = parser.parse_args(url.split())
    r = requests.get(args.url)
    soup = BeautifulSoup(r.content, 'html.parser')
    table = soup.find_all('script',id="evaluation-table-data")
    table = table[0].next
    table = table[1:-1]
    result = split_ex(table)
    if len(result) > 0:
        print(args.url+'...\n取得完了')
    dics = []

    filename = "paperwithcode_" + args.url.split('/')[-1] + ".xlsx"
    #ワークブックの作成
    wb = Workbook()
    #ワークシートの作成。第1引数にシート名、第2引数に挿入位置
    ws_table = wb.create_sheet("sota table", 0)
    metrics = {}
    for i,row in enumerate(result):
        dic = json.loads(row)
        dics.append(dic)
        for j,m in enumerate(dic["metrics"].keys()):
            metrics[m] = 0
        
    ws_table.cell(row = 1, column = 1, value = 'rank')
    ws_table.cell(row = 1, column = 2, value = 'method')
    for j,m in enumerate(metrics.keys()):
        ws_table.cell(row = 1, column = 3 + j, value = m)
    ws_table.cell(row = 1, column = 4 + j, value = 'date')
    ws_table.cell(row = 1, column = 5 + j, value = 'url')   
    
    for i,dic in enumerate(dics):  
        ws_table.cell(row = i + 2, column = 1, value = int(dic['rank']))
        ws_table.cell(row = i + 2, column = 2, value = dic['method'])
        for j,m in enumerate(dic["metrics"].keys()):
            
            if not dic['metrics'][m] == '' and not dic['metrics'][m] == None:
                ws_table.cell(row = i + 2, column = 3 + j, value = float(dic['raw_metrics'][m]))
        d = dic['evaluation_date'].split('-')
        ws_table.cell(row = i + 2, column = 4 + j, value = datetime.date(int(d[0]),int(d[1]),int(d[2]))).number_format = 'mmm\'y'
        if dic['paper']['url'] is not None:
            ws_table.cell(row = i + 2, column = 5 + j, value = 'https://paperswithcode.com' + dic['paper']['url'])
        
    ws_date = wb.create_sheet("sota date", 1)
    sota_table = sorted(dics,key=lambda x:x['evaluation_date'])
    for j,m in enumerate(sota_table[0]["metrics"].keys()):
        ws_date.cell(row = 1,column=1 + j * 4,value=m)
        ws_date.cell(row = 2,column=1 + j * 4,value='method')
        ws_date.cell(row = 2,column=2 + j * 4,value='metrics')
        ws_date.cell(row = 2,column=3 + j * 4,value='date')
        max_metric = 0
        cnt = 0
        for i,dic in enumerate(sota_table):  
            if not dic['metrics'][m] == '' and not dic['metrics'][m] == None:
                value = float(dic['raw_metrics'][m])
                if value > max_metric:
                    ws_date.cell(row = 3+cnt,column=1 + j * 4,value=dic['method'])
                    ws_date.cell(row = 3+cnt,column=2 + j * 4,value=value)
                    d = dic['evaluation_date'].split('-')
                    ws_date.cell(row = 3+cnt,column=3 + j * 4,value=datetime.date(int(d[0]),int(d[1]),int(d[2]))).number_format = 'mmm\'y'
                    max_metric = value
                    cnt+=1

    

    max_row = i+2
    x_col = 4 + j
    for k,metric in enumerate(metrics.keys()):
        ws_graph = wb.create_chartsheet(metric, 2+k)
        #ScatterChartオブジェクトを作成
        chart = ScatterChart()
        

        y_col = 3 + k
        #グラフのX軸の範囲を設定する為に、Referenceオブジェクト作る
        x_values = Reference(ws_table, min_col = x_col, min_row = 2, max_row = max_row)

        #データの書き込み
        #データの範囲(Y軸)をReferenceで選択
        values = Reference(ws_table, min_col = y_col, min_row = 2, max_row = max_row)
        #Seriesオブジェクトを作成
        series = Series(values, x_values, title=metric)
        chart.append(series)
        char_prop = CharacterProperties(latin=Font(typeface="MSP ゴシック"), sz=10 * 100, b=False) # Not bold
        para_prop = ParagraphProperties(defRPr=char_prop)
        rich_text = RichText(p=[Paragraph(pPr=para_prop, endParaRPr=char_prop)])
        if args.all_label:
            sota_names = [t[0].value for t in [s for s in ws_date.iter_rows(min_col=1+4*k, min_row=3, max_col=2+4*k, max_row=ws_table.max_row)]]
            lb = []
            for i in range(ws_table.max_row):
                if not ws_table.cell(column=2+4*k,row=2+i).value in sota_names:#SoTAのラベルと被らないようにする
                    lb.append(DataLabel(idx=i,showVal=True,numFmt=' ',separator=ws_table.cell(column=2+4*k,row=2+i).value,showCatName=True,txPr=rich_text,dLblPos='r'))
            lbl = DataLabelList(lb)
            chart.series[0].labels = lbl
        
        # ラインを無効
        chart.series[0].graphicalProperties.line.noFill = True
        # 以下マーカーの設定
        chart.series[0].marker.symbol = 'circle' # シンボルを指定
        chart.series[0].marker.size = 5                 # サイズを指定

        #SOTA推移
        sota_x_values = Reference(ws_date, min_col = 3+4*k, min_row = 3, max_row = ws_date.max_row)
        sota_values = Reference(ws_date, min_col = 2+4*k, min_row = 3, max_row = ws_date.max_row)
        series = Series(sota_values, sota_x_values, title='')
        chart.append(series)

        char_prop = CharacterProperties(latin=Font(typeface="MSP ゴシック"), sz=16 * 100, b=False) # Not bold
        para_prop = ParagraphProperties(defRPr=char_prop)
        rich_text = RichText(p=[Paragraph(pPr=para_prop, endParaRPr=char_prop)])
        lb = []
        for i in range(ws_date.max_row):
            lb.append(DataLabel(idx=i,showVal=True,numFmt=' ',separator=ws_date.cell(column=1+4*k,row=3+i).value,showSerName=True,txPr=rich_text,dLblPos='l'))
        lbl = DataLabelList(lb)
        # 以下マーカーの設定
        chart.series[1].marker.symbol = 'circle' # シンボルを指定
        chart.series[1].marker.size = 5                 # サイズを指定
        chart.series[1].labels = lbl
        
        chart.y_axis.title = metric
        chart.y_axis.title.tx.rich.p[0].pPr =  ParagraphProperties(defRPr=CharacterProperties(latin=Font(typeface="MSP ゴシック"), sz=16*100, b=True,solidFill="000000"))
        chart.x_axis.majorGridLines = None
        chart.legend = None

        ws_graph.add_chart(chart)
    print('Output:{}\nfinish!!'.format(filename))
    wb.save(filename)

if __name__ == '__main__':
    main()